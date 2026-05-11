import os
import sqlite3
from datetime import datetime, date
from io import BytesIO

import psycopg2
from flask import Flask, request, redirect, render_template, jsonify, send_file, abort
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)

ACCESS_CODE = os.environ.get("ACCESS_CODE", "h2otech")
DATABASE_URL = os.environ.get("DATABASE_URL")
LOCAL_DB = "interventions.db"


def is_postgres():
    return bool(DATABASE_URL and DATABASE_URL.startswith(("postgres://", "postgresql://")))


def get_conn():
    if is_postgres():
        return psycopg2.connect(DATABASE_URL)
    return sqlite3.connect(LOCAL_DB)


def ph():
    return "%s" if is_postgres() else "?"


def init_db():
    conn = get_conn()
    cur = conn.cursor()
    if is_postgres():
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS interventions (
                id SERIAL PRIMARY KEY,
                client TEXT NOT NULL,
                adresse TEXT NOT NULL,
                nature TEXT NOT NULL,
                urgence TEXT NOT NULL DEFAULT 'NORMALE',
                statut TEXT NOT NULL DEFAULT 'A FAIRE',
                date_intervention DATE NOT NULL,
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                done_at TIMESTAMP NULL
            )
            """
        )
    else:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS interventions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                client TEXT NOT NULL,
                adresse TEXT NOT NULL,
                nature TEXT NOT NULL,
                urgence TEXT NOT NULL DEFAULT 'NORMALE',
                statut TEXT NOT NULL DEFAULT 'A FAIRE',
                date_intervention TEXT NOT NULL,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                done_at TEXT NULL
            )
            """
        )
    conn.commit()
    conn.close()


def require_code():
    code = request.values.get("code", "")
    if code != ACCESS_CODE:
        abort(403)


def rows_to_dicts(cur):
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]


def month_bounds(target=None):
    target = target or date.today()
    start = date(target.year, target.month, 1)
    if target.month == 12:
        end = date(target.year + 1, 1, 1)
    else:
        end = date(target.year, target.month + 1, 1)
    return start, end


def get_interventions(include_done=True, month_start=None, month_end=None):
    conn = get_conn()
    cur = conn.cursor()
    params = []
    where = []

    if not include_done:
        where.append("statut <> 'TERMINE'")
    if month_start and month_end:
        where.append(f"date_intervention >= {ph()}")
        params.append(month_start)
        where.append(f"date_intervention < {ph()}")
        params.append(month_end)

    sql = """
        SELECT id, client, adresse, nature, urgence, statut,
               date_intervention, created_at, done_at
        FROM interventions
    """
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += """
        ORDER BY
            CASE WHEN urgence='URGENT' THEN 0 ELSE 1 END,
            CASE WHEN statut='A FAIRE' THEN 0 WHEN statut='EN COURS' THEN 1 ELSE 2 END,
            date_intervention ASC,
            created_at DESC
    """
    cur.execute(sql, params)
    data = rows_to_dicts(cur)
    conn.close()
    return data


@app.route("/", methods=["GET", "POST"])
def index():
    code = request.values.get("code", "")
    if code != ACCESS_CODE:
        return render_template("login.html")

    if request.method == "POST":
        client = request.form.get("client", "").strip()
        adresse = request.form.get("adresse", "").strip()
        nature = request.form.get("nature", "").strip()
        urgence = request.form.get("urgence", "NORMALE").strip().upper()
        statut = request.form.get("statut", "A FAIRE").strip().upper()
        date_intervention = request.form.get("date_intervention", "").strip() or date.today().isoformat()

        if urgence not in ["NORMALE", "URGENT"]:
            urgence = "NORMALE"
        if statut not in ["A FAIRE", "EN COURS", "TERMINE"]:
            statut = "A FAIRE"

        if client and adresse and nature:
            conn = get_conn()
            cur = conn.cursor()
            p = ph()
            cur.execute(
                f"""
                INSERT INTO interventions
                (client, adresse, nature, urgence, statut, date_intervention, done_at)
                VALUES ({p}, {p}, {p}, {p}, {p}, {p}, {p})
                """,
                (
                    client,
                    adresse,
                    nature,
                    urgence,
                    statut,
                    date_intervention,
                    datetime.now().isoformat(timespec="seconds") if statut == "TERMINE" else None,
                ),
            )
            conn.commit()
            conn.close()
        return redirect(f"/?code={ACCESS_CODE}")

    interventions = get_interventions(include_done=True)
    return render_template("index.html", interventions=interventions, code=ACCESS_CODE)


@app.route("/api/interventions")
def api_interventions():
    return jsonify(get_interventions(include_done=True))


@app.route("/ecran")
def ecran():
    return render_template("ecran.html")


@app.route("/statut/<int:intervention_id>/<statut>", methods=["POST", "GET"])
def statut(intervention_id, statut):
    require_code()
    statut = statut.upper().replace("_", " ")
    if statut not in ["A FAIRE", "EN COURS", "TERMINE"]:
        abort(400)
    done_at = datetime.now().isoformat(timespec="seconds") if statut == "TERMINE" else None
    conn = get_conn()
    cur = conn.cursor()
    p = ph()
    cur.execute(f"UPDATE interventions SET statut={p}, done_at={p} WHERE id={p}", (statut, done_at, intervention_id))
    conn.commit()
    conn.close()
    return redirect(f"/?code={ACCESS_CODE}")


@app.route("/delete/<int:intervention_id>", methods=["POST", "GET"])
def delete(intervention_id):
    require_code()
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"DELETE FROM interventions WHERE id={ph()}", (intervention_id,))
    conn.commit()
    conn.close()
    return redirect(f"/?code={ACCESS_CODE}")


@app.route("/historique")
def historique():
    require_code()
    interventions = get_interventions(include_done=True)
    return render_template("historique.html", interventions=interventions, code=ACCESS_CODE)


def build_excel(rows, title="Suivi interventions"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Interventions"
    ws.append([title])
    ws.merge_cells("A1:I1")
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Date prévue", "Client", "Adresse", "Nature intervention", "Urgence", "Statut", "Créé le", "Terminé le", "ID"]
    ws.append(headers)
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append([
            r.get("date_intervention"),
            r.get("client"),
            r.get("adresse"),
            r.get("nature"),
            r.get("urgence"),
            r.get("statut"),
            r.get("created_at"),
            r.get("done_at"),
            r.get("id"),
        ])

    widths = [16, 28, 45, 80, 16, 18, 22, 22, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    for row in ws.iter_rows(min_row=3):
        urgence = row[4].value
        statut = row[5].value
        if statut == "TERMINE":
            fill_color = "C6EFCE"
        elif urgence == "URGENT":
            fill_color = "FFC7CE"
        elif statut == "EN COURS":
            fill_color = "FFEB9C"
        else:
            fill_color = "FFFFFF"
        fill = PatternFill("solid", fgColor=fill_color)
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = fill

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


@app.route("/export/monthly")
def export_monthly():
    require_code()
    start, end = month_bounds()
    rows = get_interventions(include_done=True, month_start=start.isoformat(), month_end=end.isoformat())
    bio = build_excel(rows, f"Interventions - {start.strftime('%m/%Y')}")
    return send_file(
        bio,
        as_attachment=True,
        download_name=f"interventions_{start.strftime('%Y_%m')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/export/all")
def export_all():
    require_code()
    rows = get_interventions(include_done=True)
    bio = build_excel(rows, "Historique complet interventions")
    return send_file(
        bio,
        as_attachment=True,
        download_name="interventions_historique_complet.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


init_db()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=False)
